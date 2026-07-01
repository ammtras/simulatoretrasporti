from django.shortcuts import render, redirect
from .forms import SpedizioneForm, PaccoFormSet
from .models import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView
from django.utils import timezone
from decimal import Decimal
from django.db.models import Q
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import timedelta
from trasporti.services.calcolatore import CalcolatriceService
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import re

def check_supplemento_applicable(ids_list, spedizione):
    if sup.tipo_servizio.codice == "CONTR":
        return spedizione.da_nazione == "IT" and spedizione.a_nazione == "IT"
    return True

def aggiungi_supplementi_automatici_per_zona(ids_list, spedizione, zona_spedizioniere):
    ids_finali = list(ids_list)

    codici_da_aggiungere = []

    if spedizione.assicurazione_euro and spedizione.assicurazione_euro > 0:
        codici_da_aggiungere.append("ASSIC")

    if spedizione.contrassegno_euro and spedizione.contrassegno_euro > 0:
        codici_da_aggiungere.append("CONTR")

    codici_da_aggiungere.append("PEAKS")

    for codice in codici_da_aggiungere:
        supplementi = Supplemento.objects.filter(
            tipo_servizio__codice=codice,
            zone_tariffazione=zona_spedizioniere
        )

        if codice == "PEAKS":
            supplementi = supplementi.filter(
                valid_from__lte=spedizione.data,
                valid_to__gte=spedizione.data
            )

        for supp in supplementi:
            if supp.id not in ids_finali:
                ids_finali.append(supp.id)

    return ids_finali

def simula_preventivi(spedizione, pacchi):
    preventivi = []

    ids_supp = getattr(spedizione, "_supplementi_simulati", [])
    print(f"servizi_scelti_per_check {ids_supp}")

    zone_tratta = [
        spedizione.da_zona,
        spedizione.a_zona,
    ]

    spedizionieri = Spedizioniere.objects.all().order_by("-nome")
    # INIZIO LE MODIFICHE ORA

    for spedizioniere in spedizionieri:

        zone_candidate = Zona_spedizioniere.objects.filter(
            spedizioniere=spedizioniere,
            zona__in=zone_tratta
        ).distinct().select_related("spedizioniere")

        if not zone_candidate.exists():
            continue

        priorita_massima = max(
            z.priorita for z in zone_candidate
        )

        zone_prioritarie = [
            z for z in zone_candidate
            if z.priorita == priorita_massima
        ]

        for z_spedizioniere in zone_prioritarie:

            '''print(
                f"Spedizioniere: {spedizioniere.nome} | "
                f"Zona tratta: {zone_tratta} | "
                f"Tariffa: {z_spedizioniere.nome} | "
                f"Priorità: {z_spedizioniere.priorita}"
            )'''

            dettaglio_pesi = CalcolatriceService.dettaglio_calcolo_preventivo(
                pacchi,
                z_spedizioniere
            )
            print(f'DDDEEEEBBBUUUGGG dettaglio_pesi {dettaglio_pesi}')

            ids_supp_zona = aggiungi_supplementi_automatici_per_zona(
                ids_supp,
                spedizione,
                z_spedizioniere
            )

            context_tariffa = {
                "spedizione": spedizione,
                "pacchi": pacchi,
                "zona": z_spedizioniere,
                "peso_tassabile": dettaglio_pesi["peso_tassabile"],
                "dettaglio": dettaglio_pesi,
                "ids_supplementi": ids_supp_zona,
            }

            '''nome_spedizioniere = spedizioniere.nome.upper().strip()

            if nome_spedizioniere == "GLS":
                result_reale = CalcolatriceService._scaglioni(
                    context_tariffa
                )

            elif nome_spedizioniere == "FEDEX":
                result_reale = CalcolatriceService._a_collo(
                    context_tariffa
                )

            else:
                continue'''



            tipo_tariffazione = spedizioniere.tipo_tariffazione

            print("SPEDIZIONIERE:", z_spedizioniere.spedizioniere.nome)
            print("TIPO:", tipo_tariffazione)

            if tipo_tariffazione == Spedizioniere.A_SCAGLIONI:

                print("=== STO CHIAMANDO DAVVERO _SCAGLIONI ===")
                print("context_tariffa:", context_tariffa)
                print("SPEDIZIONIERE:", z_spedizioniere.spedizioniere.nome)
                print("ZONA:", z_spedizioniere)
                print("PESO TASSABILE PRIMA:", context_tariffa["peso_tassabile"])
                print("DETTAGLIO:", context_tariffa["dettaglio"])

                result_reale = CalcolatriceService._scaglioni(context_tariffa)

            elif tipo_tariffazione == Spedizioniere.A_COLLO:

                print("=== STO CHIAMANDO DAVVERO _A_COLLO ===")
                print("context_tariffa:", context_tariffa)
                print("peso_tassabile:", context_tariffa["peso_tassabile"])

                result_reale = CalcolatriceService._a_collo(context_tariffa)

            else:
                continue

            if result_reale:
                preventivi.append({
                    "zona_tariffazione_id": z_spedizioniere.id,
                    "zona_tariffazione_nome": z_spedizioniere.nome,
                    "zone_tratta": str(zone_tratta),
                    "trasportatore": spedizioniere.nome,
                    "totale_con_iva_euro": result_reale["totale_con_iva_euro"],
                    "dettaglio": result_reale["dettaglio"],
                })

    if preventivi:
        min_price = min(p["totale_con_iva_euro"] for p in preventivi)

        for p in preventivi:
            p["best"] = p["totale_con_iva_euro"] == min_price

    return preventivi

def loggin(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request,username=username,password=password)
        #print('a')
        if user is not None:
            #print('b')
            login(request,user)
            return redirect('crea_spedizione')
            #print('b2')
        else:
            #print('c')
            messages.success(request, ('Opppsss, qualcosa è andato storto'))
            return redirect('login')
    else:
        #print('url aperto')
        return render(request, 'trasporti/login.html', {})

@login_required
def loggout(request):
    if request.method == "POST":
        logout(request)
        return redirect('login')


def aggiungi_supplementi_automatici_per_zona(ids_list, spedizione, zona_spedizioniere):
    ids_finali = list(ids_list)

    codici_da_aggiungere = []

    if spedizione.assicurazione_euro and spedizione.assicurazione_euro > 0:
        codici_da_aggiungere.append("ASSIC")

    if spedizione.contrassegno_euro and spedizione.contrassegno_euro > 0:
        codici_da_aggiungere.append("CONTR")

    codici_da_aggiungere.append("PEAKS")

    for codice in codici_da_aggiungere:
        supplementi = Supplemento.objects.filter(
            tipo_servizio__codice=codice,
            zone_tariffazione=zona_spedizioniere
        )

        if codice == "PEAKS":
            supplementi = supplementi.filter(
                valid_from__lte=spedizione.data,
                valid_to__gte=spedizione.data
            )

        for supp in supplementi:
            if supp.id not in ids_finali:
                ids_finali.append(supp.id)

    return ids_finali

@login_required
def crea_spedizione(request):
    title = "Simulatore"
    preventivi = None

    supplementi_disponibili = Supplemento.objects.all()

    ids_supplementi = []

    def aggiungi_assic_contr_peaks(ids_list, valore_merce, valore_contr, spedizione):
        return ids_list


    if request.method == "POST":
        #print("========== POST RAW ==========")
        #print(request.POST)

        action = request.POST.get("action")
        form = SpedizioneForm(request.POST)
        formset = PaccoFormSet(request.POST)

        ids_supplementi = request.POST.getlist("supplementi_selezionati")

        if form.is_valid() and formset.is_valid():

            pacchi_data = formset.cleaned_data
            pacchi = [p for p in pacchi_data if p and not p.get("DELETE", False)]

            ids_int = [int(i) for i in ids_supplementi if i.isdigit()]

            valore_merce = form.cleaned_data.get('assicurazione_euro', Decimal('0'))
            valore_contr = form.cleaned_data.get('contrassegno_euro', Decimal('0'))

            servizi_scelti = form.cleaned_data.get("servizi_richiesti", [])

            if action == "simulate":
                spedizione_temp = form.save(commit=False)

                spedizione_temp.valore_merce = valore_merce
                spedizione_temp.valore_contrassegno = valore_contr

                ids_simulati = aggiungi_assic_contr_peaks(
                    ids_int.copy(),
                    valore_merce,
                    valore_contr,
                    spedizione_temp
                )

                spedizione_temp._servizi_simulati = list(servizi_scelti)
                print(f'servizi_scelti:{servizi_scelti}') #qui manca il fuel

                spedizione_temp._supplementi_simulati = ids_simulati


                supp_richiesti = Supplemento.objects.filter(id__in=ids_simulati)
                print('supp_richiesti')
                for s in supp_richiesti:
                    print(f"{s.id} - {s.nome}")

                preventivi = simula_preventivi(spedizione_temp, pacchi)

            elif action == "confirm":
                spedizione = form.save(commit=False)

                peso_reale_totale = sum(Decimal(str(p["peso_kg"])) for p in pacchi)
                spedizione.peso_reale_totale_kg = peso_reale_totale

                nome_trasportatore = request.POST.get("trasportatore")

                try:
                    trasportatore_obj = Spedizioniere.objects.get(nome=nome_trasportatore)
                    spedizione.trasportatore_scelto = trasportatore_obj

                    zona_corriere_obj = Zona_spedizioniere.objects.filter(
                        spedizioniere=trasportatore_obj,
                        zona__in=[spedizione.da_zona, spedizione.a_zona]
                    ).order_by('-priorita').first()

                    spedizione.zona_tariffazione_spedizioniere = zona_corriere_obj

                except Spedizioniere.DoesNotExist:
                    spedizione.trasportatore_scelto = None

                #prezzo = request.POST.get("prezzo", "0")
                totale_con_iva_euro = request.POST.get("totale_con_iva_euro", "0")
                #print("xxxx PREzzo", repr(prezzo))
                spedizione.valore_preventivo = Decimal(str(totale_con_iva_euro).replace(",", "."))


                spedizione.save()

                ids_finali = aggiungi_supplementi_automatici_per_zona(
                    ids_int,
                    spedizione,
                    zona_corriere_obj
                )

                print("=== DEBUG SALVATAGGIO SUPPLEMENTI ===")
                print("zona_corriere_obj:", zona_corriere_obj)
                print("ids_int manuali:", ids_int)
                print("assicurazione_euro:", spedizione.assicurazione_euro)
                print("contrassegno_euro:", spedizione.contrassegno_euro)
                print("ids_finali:", ids_finali)

                for s in Supplemento.objects.filter(id__in=ids_finali):
                    print(
                        f"id={s.id} | "
                        f"nome={s.nome} | "
                        f"codice={s.tipo_servizio.codice if s.tipo_servizio else None}"
                    )

                spedizione.supplementi.set(ids_finali)

                print("SUPPLEMENTI SALVATI:")
                for s in spedizione.supplementi.all():
                    print(
                        f"id={s.id} | "
                        f"nome={s.nome} | "
                        f"codice={s.tipo_servizio.codice if s.tipo_servizio else None}"
                    )

                formset.instance = spedizione
                formset.save()


                return redirect("spedizioni")

    else:
        form = SpedizioneForm(instance=Spedizione(data=timezone.now().date()))
        formset = PaccoFormSet()

    return render(request, "trasporti/crea_spedizione.html", {
        "title":title,
        "form": form,
        "formset": formset,
        "preventivi": preventivi,
        "supplementi_disponibili": supplementi_disponibili,
        "ids_selezionati": [i for i in ids_supplementi if i.isdigit()]
    })

class spedizioni(LoginRequiredMixin,ListView):
    model = Spedizione
    template_name = "trasporti/spedizioni.html"
    context_object_name = "spedizioni"
    ordering = ["-data", "-id"]




    def get_paginate_by(self, queryset):
        valori_validi = [4, 5, 10, 30, 50, 100]

        profilo, created = Profilo.objects.get_or_create(
            user=self.request.user
        )

        per_page = self.request.GET.get("per_page")

        if per_page:
            try:
                per_page = int(per_page)
            except ValueError:
                per_page = profilo.righe_per_pagina

            if per_page in valori_validi:
                profilo.righe_per_pagina = per_page
                profilo.save(update_fields=["righe_per_pagina"])
                return per_page

        return profilo.righe_per_pagina



    def get_queryset(self):
        # 1. Ottieni il queryset base
        queryset = super().get_queryset().prefetch_related('pacchi')

        # 2. Leggi i parametri dal form
        q = self.request.GET.get('q')
        periodo = self.request.GET.get("periodo")
        oggi = timezone.localdate()
        data_da = self.request.GET.get('data_da')
        data_a = self.request.GET.get('data_a')



        # 3. Filtra il queryset
        if q:
            # Filtra per città di partenza, arrivo o nome trasportatore
            queryset = queryset.filter(
                Q(da_cliente_citta__icontains=q) |
                Q(a_cliente_citta__icontains=q)
            )



        if periodo == "oggi":
            data_da = oggi
            data_a = oggi

        elif periodo == "ieri":
            ieri = oggi - timedelta(days=1)
            data_da = ieri
            data_a = ieri

        elif periodo == "settimana":
            data_da = oggi - timedelta(days=oggi.weekday())
            data_a = oggi

        elif periodo == "settimana_scorsa":
            inizio_settimana_corrente = oggi - timedelta(days=oggi.weekday())
            data_da = inizio_settimana_corrente - timedelta(days=7)
            data_a = inizio_settimana_corrente - timedelta(days=1)

        elif periodo == "mese":
            data_da = oggi.replace(day=1)
            data_a = oggi

        elif periodo == "mese_scorso":
            primo_giorno_mese_corrente = oggi.replace(day=1)
            ultimo_giorno_mese_scorso = primo_giorno_mese_corrente - timedelta(days=1)
            data_da = ultimo_giorno_mese_scorso.replace(day=1)
            data_a = ultimo_giorno_mese_scorso

        elif periodo == "anno":
            data_da = oggi.replace(month=1, day=1)
            data_a = oggi

        if data_da:
            queryset = queryset.filter(data__gte=data_da)

        if data_a:
            queryset = queryset.filter(data__lte=data_a)


        trasportatore = self.request.GET.get('trasportatore')
        # Filtro più robusto
        if trasportatore and trasportatore.isdigit():  # Controlla che sia un numero valido
            queryset = queryset.filter(trasportatore_scelto__id=int(trasportatore))


        return queryset

    def get_context_data(self, **kwargs):
        # Chiama il super per ottenere il queryset già filtrato!
        context = super().get_context_data(**kwargs)

        context["title"] = "Spedizioni"

        context['trasportatori'] = Spedizioniere.objects.all().order_by('nome')

        # 3. Il tuo codice per il calcolo dinamico (CalcolatriceService)
        # Ora il ciclo opererà SOLO sugli elementi filtrati
        for s in context["spedizioni"]:
            #if s.trasportatore_scelto and s.trasportatore_scelto.nome.upper() == "GLS":
            if s.trasportatore_scelto and s.zona_tariffazione_spedizioniere:
                # ... (il resto del tuo codice di calcolo resta invariato)
                pacchi_list = [
                    {"altezza_cm": p.altezza_cm, "larghezza_cm": p.larghezza_cm, "profondita_cm": p.profondita_cm,
                     "peso_kg": p.peso_kg}
                    for p in s.pacchi.all()
                ]
                result = CalcolatriceService.calcola(s, pacchi_list)
                if result:
                    s.dettaglio = result.get("dettaglio", {})
                    s.valore_preventivo_aggiornato = result.get("totale_con_iva_euro")

        return context


#vedere di rinominare, qui è un "get zona_spedizioniere e supplementi"
@login_required
def controllo_tariffe(request):
    title = "Controllo Tariffe"
    # Carichiamo tutto in un colpo solo per efficienza

    spedizionieri = Spedizioniere.objects.prefetch_related(
        'sspedizioniere__zona_spedizioniere',
        'sspedizioniere__supplementi',
    ).all()

    return render(request, 'trasporti/controllo_tariffe.html', {'spedizionieri': spedizionieri,'title':title})

def estrai_decimal_da_stringa(value):
    value = str(value)
    value = re.sub(r"<[^>]*>", "", value)
    value = value.replace("€", "").replace("kg", "").replace("%", "")
    value = value.replace(",", ".")
    numeri = re.findall(r"\d+(?:\.\d+)?", value)

    if numeri:
        return Decimal(numeri[-1])

    return Decimal("0")

#export per controllo fatture GLS
@login_required
def xxesporta_spedizioni_excel(request):
    view = spedizioni()
    view.request = request

    #queryset = view.get_queryset().prefetch_related("pacchi", "supplementi")
    queryset = (view.get_queryset().order_by("data", "id").prefetch_related("pacchi", "supplementi"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Spedizioni"

    riga = 1


    headers = [
        "Data",
        "Cliente da",
        "Cliente a",
        "Totale colli",
        "Peso reale kg",
        "Peso volume kg",
        "Nolo €",
        "Contrassegno €",
        "Assicurazione €",
        "Fuel €",
        "Fuel %",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(riga, col, header).font = Font(bold=True)

    riga += 1

    # Memorizza la prima riga dei dati
    prima_riga_dati = riga

    for s in queryset:
        pacchi_list = [
            {
                "altezza_cm": p.altezza_cm,
                "larghezza_cm": p.larghezza_cm,
                "profondita_cm": p.profondita_cm,
                "peso_kg": p.peso_kg,
            }
            for p in s.pacchi.all()
        ]

        result = None

        if s.trasportatore_scelto and s.zona_tariffazione_spedizioniere:
            result = CalcolatriceService.calcola(s, pacchi_list)

        dettaglio_items = (
            result.get("dettaglio", {}).get("items", [])
            if result
            else []
        )

        peso_reale = sum(
            Decimal(str(p.get("peso_kg") or 0))
            for p in pacchi_list
        )

        peso_volume = Decimal("0")
        nolo = Decimal("0")
        costo_contrassegno = Decimal("0")
        costo_assicurazione = Decimal("0")
        fuel_euro = Decimal("0")
        fuel_percentuale = ""

        for item in dettaglio_items:
            label = str(item.get("label", "")).lower()
            value = str(item.get("value", ""))

            if "peso volume" in label:
                peso_volume = estrai_decimal_da_stringa(value)

            elif "nolo" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "scaglione" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "supplementi" in label:
                righe = value.split("<br>")

                for r in righe:
                    r_lower = r.lower()

                    if "contr" in r_lower:
                        costo_contrassegno += estrai_decimal_da_stringa(r)

                    elif "assic" in r_lower or "assicurazione" in r_lower:
                        costo_assicurazione += estrai_decimal_da_stringa(r)

                    elif "fuel" in r_lower:
                        fuel_euro += estrai_decimal_da_stringa(r)

                        match = re.search(r"\(([\d.,]+)%\)", r)
                        if match:
                            fuel_percentuale = match.group(1).replace(",", ".")

        ws.cell(riga, 1, s.data)
        ws.cell(riga, 2, s.da_cliente_citta)
        ws.cell(riga, 3, s.a_cliente_citta)
        ws.cell(riga, 4, s.pacchi.count())
        ws.cell(riga, 5, float(peso_reale))
        ws.cell(riga, 6, float(peso_volume))
        ws.cell(riga, 7, float(nolo))
        ws.cell(riga, 8, float(costo_contrassegno))
        ws.cell(riga, 9, float(costo_assicurazione))
        ws.cell(riga, 10, float(fuel_euro))
        ws.cell(riga, 11, fuel_percentuale)

        riga += 1

    ultima_riga_dati = riga - 1
    ws.cell(riga, 1, f"Spedizioni: {queryset.count()}").font = Font(bold=True)
    ws.cell(riga, 6, "TOTALI").font = Font(bold=True)
    ws.cell(riga, 7, f"=SUM(G{prima_riga_dati}:G{ultima_riga_dati})")
    ws.cell(riga, 8, f"=SUM(H{prima_riga_dati}:H{ultima_riga_dati})")
    ws.cell(riga, 9, f"=SUM(I{prima_riga_dati}:I{ultima_riga_dati})")
    ws.cell(riga, 10, f"=SUM(J{prima_riga_dati}:J{ultima_riga_dati})")

    for c in range(6, 11):
        ws.cell(riga, c).font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="spedizioni.xlsx"'

    wb.save(response)
    return response

def YYesporta_spedizioni_excel(request):
    view = spedizioni()
    view.request = request

    queryset = (
        view.get_queryset()
        .order_by("data", "id")
        .prefetch_related("pacchi", "supplementi")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Spedizioni"

    riga = 1

    headers = [
        "Data",
        "Cliente da",
        "Cliente a",
        "Totale colli",
        "Peso reale kg",
        "Peso volume kg",
        "Nolo €",
        "Contrassegno €",
        "Assicurazione €",
        "Fuel €",
        "Fuel %",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(riga, col, header).font = Font(bold=True)

    riga += 1
    prima_riga_dati = riga

    riepilogo_per_data = {}

    for s in queryset:
        pacchi_list = [
            {
                "altezza_cm": p.altezza_cm,
                "larghezza_cm": p.larghezza_cm,
                "profondita_cm": p.profondita_cm,
                "peso_kg": p.peso_kg,
            }
            for p in s.pacchi.all()
        ]

        result = None

        if s.trasportatore_scelto and s.zona_tariffazione_spedizioniere:
            result = CalcolatriceService.calcola(s, pacchi_list)

        dettaglio_items = (
            result.get("dettaglio", {}).get("items", [])
            if result
            else []
        )

        peso_reale = sum(
            Decimal(str(p.get("peso_kg") or 0))
            for p in pacchi_list
        )

        peso_volume = Decimal("0")
        nolo = Decimal("0")
        costo_contrassegno = Decimal("0")
        costo_assicurazione = Decimal("0")
        fuel_euro = Decimal("0")
        fuel_percentuale = ""

        for item in dettaglio_items:
            label = str(item.get("label", "")).lower()
            value = str(item.get("value", ""))

            if "peso volume" in label:
                peso_volume = estrai_decimal_da_stringa(value)

            elif "nolo" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "scaglione" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "supplementi" in label:
                righe = value.split("<br>")

                for r in righe:
                    r_lower = r.lower()

                    if "contr" in r_lower:
                        costo_contrassegno += estrai_decimal_da_stringa(r)

                    elif "assic" in r_lower or "assicurazione" in r_lower:
                        costo_assicurazione += estrai_decimal_da_stringa(r)

                    elif "fuel" in r_lower:
                        fuel_euro += estrai_decimal_da_stringa(r)

                        match = re.search(r"\(([\d.,]+)%\)", r)
                        if match:
                            fuel_percentuale = match.group(1).replace(",", ".")

        ws.cell(riga, 1, s.data)
        ws.cell(riga, 2, s.da_cliente_citta)
        ws.cell(riga, 3, s.a_cliente_citta)
        ws.cell(riga, 4, s.pacchi.count())
        ws.cell(riga, 5, float(peso_reale))
        ws.cell(riga, 6, float(peso_volume))
        ws.cell(riga, 7, float(nolo))
        ws.cell(riga, 8, float(costo_contrassegno))
        ws.cell(riga, 9, float(costo_assicurazione))
        ws.cell(riga, 10, float(fuel_euro))
        ws.cell(riga, 11, fuel_percentuale)

        data_key = s.data

        if data_key not in riepilogo_per_data:
            riepilogo_per_data[data_key] = {
                "numero_spedizioni": 0,
                "totale_colli": 0,
                "nolo": Decimal("0"),
                "contrassegno": Decimal("0"),
                "assicurazione": Decimal("0"),
                "fuel": Decimal("0"),
            }

        riepilogo_per_data[data_key]["numero_spedizioni"] += 1
        riepilogo_per_data[data_key]["totale_colli"] += s.pacchi.count()
        riepilogo_per_data[data_key]["nolo"] += nolo
        riepilogo_per_data[data_key]["contrassegno"] += costo_contrassegno
        riepilogo_per_data[data_key]["assicurazione"] += costo_assicurazione
        riepilogo_per_data[data_key]["fuel"] += fuel_euro

        riga += 1

    ultima_riga_dati = riga - 1

    ws.cell(riga, 1, f"Spedizioni: {queryset.count()}").font = Font(bold=True)
    ws.cell(riga, 6, "TOTALI").font = Font(bold=True)
    ws.cell(riga, 7, f"=SUM(G{prima_riga_dati}:G{ultima_riga_dati})")
    ws.cell(riga, 8, f"=SUM(H{prima_riga_dati}:H{ultima_riga_dati})")
    ws.cell(riga, 9, f"=SUM(I{prima_riga_dati}:I{ultima_riga_dati})")
    ws.cell(riga, 10, f"=SUM(J{prima_riga_dati}:J{ultima_riga_dati})")

    for c in range(1, 11):
        ws.cell(riga, c).font = Font(bold=True)

    riga += 3

    ws.cell(riga, 1, "Riepilogo per giorno").font = Font(bold=True)
    riga += 1

    headers_riepilogo = [
        "Data",
        "Numero spedizioni",
        "Nolo €",
        "Contrassegno €",
        "Assicurazione €",
        "Fuel €",
    ]

    for col, header in enumerate(headers_riepilogo, start=1):
        ws.cell(riga, col, header).font = Font(bold=True)

    riga += 1
    prima_riga_riepilogo = riga

    for data_key in sorted(riepilogo_per_data.keys()):
        dati = riepilogo_per_data[data_key]

        ws.cell(riga, 1, data_key)
        ws.cell(riga, 2, dati["numero_spedizioni"])
        ws.cell(riga, 3, float(dati["nolo"]))
        ws.cell(riga, 4, float(dati["contrassegno"]))
        ws.cell(riga, 5, float(dati["assicurazione"]))
        ws.cell(riga, 6, float(dati["fuel"]))

        riga += 1

    ultima_riga_riepilogo = riga - 1

    ws.cell(riga, 1, "TOTALI").font = Font(bold=True)
    ws.cell(riga, 2, f"=SUM(B{prima_riga_riepilogo}:B{ultima_riga_riepilogo})")
    ws.cell(riga, 3, f"=SUM(C{prima_riga_riepilogo}:C{ultima_riga_riepilogo})")
    ws.cell(riga, 4, f"=SUM(D{prima_riga_riepilogo}:D{ultima_riga_riepilogo})")
    ws.cell(riga, 5, f"=SUM(E{prima_riga_riepilogo}:E{ultima_riga_riepilogo})")
    ws.cell(riga, 6, f"=SUM(F{prima_riga_riepilogo}:F{ultima_riga_riepilogo})")
    ws.cell(riga, 7, f"=SUM(G{prima_riga_riepilogo}:G{ultima_riga_riepilogo})")

    for c in range(1, 10):
        ws.cell(riga, c).font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="spedizioni.xlsx"'

    wb.save(response)
    return response

def esporta_spedizioni_excel(request):
    view = spedizioni()
    view.request = request

    queryset = (
        view.get_queryset()
        .order_by("data", "id")
        .prefetch_related("pacchi", "supplementi")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Spedizioni"

    riga = 1

    headers = [
        "Data",
        "Cliente da",
        "Cliente a",
        "Totale colli",
        "Peso reale kg",
        "Peso volume kg",
        "Nolo €",
        "Contrassegno €",
        "Assicurazione €",
        "Fuel €",
        "Fuel %",
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(riga, col, header).font = Font(bold=True)

    riga += 1
    prima_riga_dati = riga

    riepilogo_per_data = {}

    for s in queryset:
        pacchi_list = [
            {
                "altezza_cm": p.altezza_cm,
                "larghezza_cm": p.larghezza_cm,
                "profondita_cm": p.profondita_cm,
                "peso_kg": p.peso_kg,
            }
            for p in s.pacchi.all()
        ]

        result = None

        if s.trasportatore_scelto and s.zona_tariffazione_spedizioniere:
            result = CalcolatriceService.calcola(s, pacchi_list)

        dettaglio_items = (
            result.get("dettaglio", {}).get("items", [])
            if result
            else []
        )

        peso_reale = sum(
            Decimal(str(p.get("peso_kg") or 0))
            for p in pacchi_list
        )

        peso_volume = Decimal("0")
        nolo = Decimal("0")
        costo_contrassegno = Decimal("0")
        costo_assicurazione = Decimal("0")
        fuel_euro = Decimal("0")
        fuel_percentuale = ""

        for item in dettaglio_items:
            label = str(item.get("label", "")).lower()
            value = str(item.get("value", ""))

            if "peso volume" in label:
                # Prende il valore prima di "kg", non il divisore
                match = re.search(r"([\d.,]+)\s*kg", value)
                if match:
                    peso_volume = Decimal(match.group(1).replace(",", "."))
                else:
                    peso_volume = estrai_decimal_da_stringa(value)

            elif "nolo" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "scaglione" in label:
                nolo = estrai_decimal_da_stringa(value)

            elif "supplementi" in label:
                righe = value.split("<br>")

                for r in righe:
                    r_lower = r.lower()

                    if "contr" in r_lower:
                        costo_contrassegno += estrai_decimal_da_stringa(r)

                    elif "assic" in r_lower or "assicurazione" in r_lower:
                        costo_assicurazione += estrai_decimal_da_stringa(r)

                    elif "fuel" in r_lower:
                        fuel_euro += estrai_decimal_da_stringa(r)

                        match = re.search(r"\(([\d.,]+)%\)", r)
                        if match:
                            fuel_percentuale = match.group(1).replace(",", ".")

        ws.cell(riga, 1, s.data)
        ws.cell(riga, 2, s.da_cliente_citta)
        ws.cell(riga, 3, s.a_cliente_citta)
        ws.cell(riga, 4, s.pacchi.count())
        ws.cell(riga, 5, float(peso_reale))
        ws.cell(riga, 6, float(peso_volume))
        ws.cell(riga, 7, float(nolo))
        ws.cell(riga, 8, float(costo_contrassegno))
        ws.cell(riga, 9, float(costo_assicurazione))
        ws.cell(riga, 10, float(fuel_euro))
        ws.cell(riga, 11, fuel_percentuale)

        data_key = s.data

        if data_key not in riepilogo_per_data:
            riepilogo_per_data[data_key] = {
                "numero_spedizioni": 0,
                "nolo": Decimal("0"),
                "contrassegno": Decimal("0"),
                "assicurazione": Decimal("0"),
                "fuel": Decimal("0"),
            }

        riepilogo_per_data[data_key]["numero_spedizioni"] += 1
        riepilogo_per_data[data_key]["nolo"] += nolo
        riepilogo_per_data[data_key]["contrassegno"] += costo_contrassegno
        riepilogo_per_data[data_key]["assicurazione"] += costo_assicurazione
        riepilogo_per_data[data_key]["fuel"] += fuel_euro

        riga += 1

    ultima_riga_dati = riga - 1

    ws.cell(riga, 1, f"Spedizioni: {queryset.count()}").font = Font(bold=True)
    ws.cell(riga, 6, "TOTALI").font = Font(bold=True)
    ws.cell(riga, 7, f"=SUM(G{prima_riga_dati}:G{ultima_riga_dati})")
    ws.cell(riga, 8, f"=SUM(H{prima_riga_dati}:H{ultima_riga_dati})")
    ws.cell(riga, 9, f"=SUM(I{prima_riga_dati}:I{ultima_riga_dati})")
    ws.cell(riga, 10, f"=SUM(J{prima_riga_dati}:J{ultima_riga_dati})")
    # TOTALONE G+H+I+J
    ws.cell(riga, 11, f"=SUM(G{riga}:J{riga})")

    for c in range(1, 12):
        ws.cell(riga, c).font = Font(bold=True)

    riga += 3

    ws.cell(riga, 1, "Riepilogo per giorno").font = Font(bold=True)
    riga += 1

    headers_riepilogo = [
        "Data",
        "Numero spedizioni",
        "Nolo €",
        "Contrassegno €",
        "Assicurazione €",
        "Fuel €",
        "Tot. imp.",
    ]

    for col, header in enumerate(headers_riepilogo, start=1):
        ws.cell(riga, col, header).font = Font(bold=True)

    riga += 1
    prima_riga_riepilogo = riga

    for data_key in sorted(riepilogo_per_data.keys()):
        dati = riepilogo_per_data[data_key]

        ws.cell(riga, 1, data_key)
        ws.cell(riga, 2, dati["numero_spedizioni"])
        ws.cell(riga, 3, float(dati["nolo"]))
        ws.cell(riga, 4, float(dati["contrassegno"]))
        ws.cell(riga, 5, float(dati["assicurazione"]))
        ws.cell(riga, 6, float(dati["fuel"]))

        riga += 1

    ultima_riga_riepilogo = riga - 1

    ws.cell(riga, 1, "TOTALI").font = Font(bold=True)
    ws.cell(riga, 2, f"=SUM(B{prima_riga_riepilogo}:B{ultima_riga_riepilogo})")
    ws.cell(riga, 3, f"=SUM(C{prima_riga_riepilogo}:C{ultima_riga_riepilogo})")
    ws.cell(riga, 4, f"=SUM(D{prima_riga_riepilogo}:D{ultima_riga_riepilogo})")
    ws.cell(riga, 5, f"=SUM(E{prima_riga_riepilogo}:E{ultima_riga_riepilogo})")
    ws.cell(riga, 6, f"=SUM(F{prima_riga_riepilogo}:F{ultima_riga_riepilogo})")
    ws.cell(riga, 7, f"=SUM(C{riga}:F{riga})")

    for c in range(1, 7):
        ws.cell(riga, c).font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="spedizioni.xlsx"'

    wb.save(response)
    return response